import time
import sys
import random
import os
from instrument import instrument_set as instr
from actions import get_data as test_item

#默认上电 VIN = 5V / 0.3A   (目前是测试阶段，仅需上电超过4V不超过5V即可)

"""
1. 该脚本是测试执行脚本, 所有测试项均在test_configs/test.xlsx中定义好测试项名称、测试参数等, 以及在test_configs/register.xlsx中定义好寄存器名称和地址
2. 每个测试项对应一个函数, 函数名可以自定义, 但建议以测试项名称开头, 方便维护和调用
3. 该脚本只是执行测试项, 该自动化测试平台需要在该项目的主目录下的main.py中启动, 通过main.py中的交互式菜单选择需要执行的测试项
4. 每个测试函数需要完成以下工作:
    4.1 从test.xlsx中读取该测试项的测试参数(如Min/Typ/Max等), 以及需要用到的寄存器地址(通过register.xlsx查询)
    4.2 根据测试需求编写测试逻辑, 包括但不限于:
        - 通过instr模块控制测试设备(如示波器、万用表等)进行测量
        - 通过instr模块控制芯片寄存器进行设置
        - 根据测量结果和测试参数判断测试是否通过
    4.3 将测试结果以excel格式或者方便整理数据的格式保存到Save_Results目录下, 方便后续分析
5. 该代码是基于test_configs目录下的OCS-QA-Q2-010-07-OCD2810  Final Test specifications V2.1.0.docx文档中的测试项编写的, 目前已经完成了其中的部分测试项, 
    后续会逐渐完善剩余测试项的自动化脚本
    该脚本不是最终版本, 目前处于测试阶段, 后续会根据测试情况不断优化和完善,如有错误地方请以上面提到的.docx文档为准, 以免造成不必要的麻烦, 


"""

#这里存放全局变量, 某个函数的某个变量需要在其他函数使用可以存入这个字典
global_use = {"LDO_TRIM":0, "LIRC_TRIM":0, "HIRC_TRIM":0, "ADC_TRIM":0, "VREF_TRIM":0, "Voffset_M":0, "Voffset_P":0}

# 基于脚本所在目录构造Excel路径, 避免CWD影响
_script_dir = os.path.dirname(os.path.abspath(__file__))
_register_path = os.path.join(_script_dir, "..", "test_configs", "register.xlsx")
_testdata_path = os.path.join(_script_dir, "..", "test_configs", "test.xlsx")

try:
    df_register = test_item.load_excel_data(_register_path)
    df_test_data = test_item.load_excel_data(_testdata_path)
except Exception as e:
    print(f"Excel数据加载失败: {e}")
    sys.exit(1)

def frange(start, stop, step=1.0):
    current = start
    while current < stop:
        yield current
        current += step

def OS_Test():
    """5.1 OS测试"""
    
def UVLO_RSTN_Test():
    """5.2 UVLO_RST自动化修调测试"""
    test_item_name_h = "UVLO_RSTH"
    test_item_name_l = "UVLO_RST_HYS"
    test_item_name_int_h = "UVLO_INT"
    test_item_name_int_l = "UVLO_INT_HYS"
    row_h = df_test_data[df_test_data["测试项"] == test_item_name_h]
    row_l = df_test_data[df_test_data["测试项"] == test_item_name_l]
    row_int_h = df_test_data[df_test_data["测试项"] == test_item_name_int_h]
    row_int_l = df_test_data[df_test_data["测试项"] == test_item_name_int_l]
    if row_h.empty and row_l.empty and row_int_h.empty and row_int_l.empty:
        print("未找到测试项")
        return
    try:
        min_val_h = float(row_h["Min."].values[0])
        max_val_h = float(row_h["Max."].values[0])

        min_val_l = float(row_l["Min."].values[0])
        max_val_l = float(row_l["Max."].values[0])

        min_val_int_h = float(row_int_h["Min."].values[0])
        max_val_int_h = float(row_int_h["Max."].values[0])

        min_val_int_l = float(row_int_l["Min."].values[0])
        max_val_int_l = float(row_int_l["Max."].values[0])

        unit = str(row_h["Unit"].values[0])
    except (IndexError, ValueError) as e:
        print(f"❌ 测试参数读取失败: {e}")
        return
    
    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register)
        RW_TEST0_SEL = test_item.get_32bit_reg_address("RW_TEST0_SEL", df_register)
        RW_TEST0_EN = test_item.get_32bit_reg_address("RW_TEST0_EN", df_register)
    except Exception as e:
        print(f"❌ 寄存器地址获取失败: {e}")
        return
    
    """
    这里没有电源选择接口了, 直接提示用户接入电源并上电即可
    后续如果有电源设备了可以通过instr模块控制电源输出, 目前先手动控制电源比较方便
    下面需要接入的也会做出提示，由于测试平台还未完善，所以有阻塞式输入，等待用户操作完成后再继续执行测试脚本
    后续等平台逐渐完善再替换成自动化脚本代码，取代手动输入
    """
    print("请将输入电源线接入芯片电源端，同时开启电流监测, 同时万用表接入P13监测电压")
    input()
    
    instr.swd_modify_register(Pad1_ana_en, 0x1, 11, 11)
    
    rst_h = 3000
    instr.serial_send(f"spi_cmd_set_val mV {rst_h}\r\n")
    for v in range(3100, 3500, 100):
        instr.serial_send(f"spi_cmd_set_val mV {v}\r\n")
        time.sleep(0.5)
        i = instr.dmm_measure_single_voltage()
        if i > 0.3 :        #这里判断电流突变的标准没有，目前先设置为0.3A, 之后可以根据测试情况调整
            rst_h = v
            break
    UVLO_RSTH = rst_h / 1000.0
    print(f"✅ 找到UVLO_RSTH = {rst_h}mV")

    if(UVLO_RSTH >= min_val_h and UVLO_RSTH <= max_val_h):
        print("UVLO_RSTH pass")
    else:
        print("UVLO_RSTH error")
    """
    这里需要需要将测试结果保存到Save_Results目录, 方便后续分析
    目前Save_Results目录下没有文件, 可以先创建一个xlsx文件放入, 之后再完善保存功能
    以下需要保存结果的注释为"保存结果"，可以在该注释下面一行写出保存结果
    的脚本
     """

    rst_l = rst_h + 500
    #这里输入电压为 rst_l
    instr.serial_send(f"spi_cmd_set_val mV {rst_l}\r\n")
    for v in range(3500, 2800, -100):
        #设置输入电压为 v
        instr.serial_send(f"spi_cmd_set_val mV {v}\r\n")
        time.sleep(0.5)
        i = instr.dmm_measure_single_current()
        if i < 0.2 :
            rst_l = v
            break
    UVLO_RST_HYS = (rst_h - rst_l) / 1000.0
    if UVLO_RST_HYS >= min_val_l and UVLO_RST_HYS <= max_val_l :
        print("UVLO_RST_HYS pass")
    else :
        print("UVLO_RST_HYS error")
    """保存结果"""

    """INT_TEST"""
    instr.swd_modify_register(RW_TEST0_SEL, 0x5, 16, 18)
    instr.swd_modify_register(RW_TEST0_EN, 0x1, 15 ,15)
    time.sleep(0.1)

    int_h = rst_l
    instr.serial_send(f"spi_cmd_set_val mV {int_h}\r\n")
    for v  in range(3100, 4500, 100):

        #设置输入电压为 v
        instr.serial_send(f"spi_cmd_set_val mV {v}\r\n")
        time.sleep(0.5)
        i = instr.dmm_measure_single_voltage()
        if i > 0.3 :
            int_h = v
            break
    UVLO_INT = int_h / 1000.0
    print(f"✅ 找到UVLO_INTH = {UVLO_INT}{unit}")
    if(UVLO_INT >= min_val_int_h and UVLO_INT <= max_val_int_h):
        print("UVLO_INT pass")
    else:
        print("UVLO_INT error")
    """保存结果"""

    int_l = int_h + 500

    #这里输入电压为 int_l
    instr.serial_send(f"spi_cmd_set_val mV {int_l}\r\n")
    time.sleep(0.5)
    for v in range(int_h + 500, 2800, -100):
        #设置输入电压为 v
        instr.serial_send(f"spi_cmd_set_val mV {v}\r\n")
        time.sleep(0.5)
        i = instr.dmm_measure_single_voltage()
        if i < 0.2 :
            int_l = v
            break
    UVLO_INT_HYS = (int_h - int_l) / 1000.0
    if UVLO_INT_HYS >= min_val_int_l and UVLO_INT_HYS <= max_val_int_l :
        print("UVLO_INT_HYS pass")
    else :
        print("UVLO_INT_HYS error")
    """保存结果"""

def OSC33K_Test():
    """5.3 OSC33K自动化修调测试"""
    test_item_name = "OSC_33K"
    row = df_test_data[df_test_data["测试项"] == test_item_name]
    if row.empty:
        print(f"未找到测试项: {test_item_name}")
        return
    try:
        min_val = float(row["Min."].values[0])
        typ_val = float(row["Typ."].values[0])
        max_val = float(row["Max."].values[0])
        unit = str(row["Unit"].values[0])
    except (IndexError, ValueError) as e:
        print(f"❌ 测试参数读取失败: {e}")
        return
    
    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register)
        RW_TEST0_SEL = test_item.get_32bit_reg_address("RW_TEST0_SEL", df_register)
        RW_TEST0_EN = test_item.get_32bit_reg_address("RW_TEST0_EN", df_register)
        RW_OSC_33K_TRIM = test_item.get_32bit_reg_address("RW_OSC_33K_TRIM", df_register)
    except Exception as e:
        print(f"❌ 寄存器地址获取失败: {e}")
        return
    
    #下面这个注释的意思是需要设置输入电压为5V, 目前没有电源设备接口, 先提示用户手动设置电压, 之后有了电源设备接口了再替换成自动化控制脚本
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")

    print("请将示波器的CH1探头接入P13引脚, 按下enter键继续...")
    input()
    time.sleep(1)
    try:
        #初始化、写入测试指令
        instr.swd_modify_register(Pad1_ana_en, 0x1, 11, 11)
        instr.swd_modify_register(RW_TEST0_SEL, 0x07, 16, 18)
        instr.swd_modify_register(RW_TEST0_EN, 0x1, 15, 15)
        time.sleep(0.2)
    except Exception as e:
        print(f"❌ 寄存器写入/万用表读取失败: {e}")
        return

    typ_val_period = 1000.0 / float(typ_val)   #测试项写的是频率, 需要转换成周期
    print(f"测试目标频率为type_val={typ_val * 1000} Hz, 目标周期为{typ_val_period:.3f} us")

    instr.osc_auto_set()
    time.sleep(2)

    instr.swd_modify_register(RW_OSC_33K_TRIM, 0x40, 16, 22)
    time.sleep(0.5)
    OSC33KA_freq, OSC33KA_period, OSC33KA_duty = instr.osc_measure_frequency("CH1")
    print(f"OSC33K频率A: {OSC33KA_freq:.3f} Hz, 周期: {OSC33KA_period:.3f} us")

    instr.swd_modify_register(RW_OSC_33K_TRIM, 0x41, 16, 22)
    time.sleep(0.5)
    OSC33KB_freq, OSC33KB_period, OSC33KB_duty = instr.osc_measure_frequency("CH1")
    print(f"OSC33K频率B: {OSC33KB_freq:.3f} Hz, 周期: {OSC33KB_period:.3f} us")

    delta_osc = OSC33KB_period - OSC33KA_period
    min_error = float('inf')

    if abs(delta_osc) < 1e-6:
        print(f"⚠️ OSC33K修调码线性插值失败(分母为0), 切换到全局遍历模式(0~127)")
        search_range = range(28, 40)
    else:
        TRIMCODE = (typ_val_period - OSC33KA_period) / delta_osc + 64
        TRIM_CALC = round(TRIMCODE)     #四舍五入取整
        search_range = range(TRIM_CALC - 4, TRIM_CALC + 6)
        print(f"OSC33K修调码计算结果: {TRIMCODE:.2f}, 四舍五入取整: {TRIM_CALC}, 搜索范围: {search_range.start} ~ {search_range.stop - 1}")

    for test_trim in search_range:
        if test_trim < 0 or test_trim > 127:
            continue
        instr.swd_modify_register(RW_OSC_33K_TRIM, test_trim, 16, 22)
        time.sleep(0.5)
        #复测，计算test_period与typ_val的绝对误差，取最小值对应的最优trim
        test_freq, test_period, test_duty = instr.osc_measure_frequency("CH1")
        test_error = abs(test_period - typ_val_period)
        print(f"测试修调码 {test_trim}: 频率={test_freq:.3f} Hz, 周期={test_period:.3f} us, 与目标周期误差={test_error:.3f} us")

        if test_error < min_error:
            min_error = test_error
            best_trim = test_trim

    instr.swd_modify_register(RW_OSC_33K_TRIM, best_trim, 16, 22)
    time.sleep(0.5)
    global_use["LIRC_TRIM"] = RW_OSC_33K_TRIM
    final_freq, final_period, final_duty = instr.osc_measure_frequency("CH1")
    print(f"最终OSC33K频率: {final_freq:.3f} Hz, 周期: {final_period:.3f} us")
    
    final_freq = final_freq / 1000      #注意示波器测试出来是的频率单位是Hz, 需要转换到KHz进行比较
    if final_freq >= min_val and final_freq <= max_val:
        print("OSC33K Test pass")
    else:
        print("OSC33K Test failed")
        print(f"修调范围为: {min_val}~{max_val} {unit}, 实际测量值: {final_freq:.3f} {unit}")
    """保存结果"""

def OSC24M_Test():
    """5.4 OSC24M自动化修调测试"""
    test_item_name = "OSC_24M"
    #获取测试数据、寄存器地址
    row = df_test_data[df_test_data["测试项"] == test_item_name]
    if row.empty:
        print(f"未找到测试项: {test_item_name}")
        return
    try:
        min_val = float(row["Min."].values[0])
        typ_val = float(row["Typ."].values[0])
        max_val = float(row["Max."].values[0])
        unit = str(row["Unit"].values[0])
        #TARGET_PERIOD = 1000.0 / typ_val
    except (IndexError, ValueError) as e:
        print(f"❌ 测试参数读取失败: {e}")
        return
    
    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register)
        RW_TEST0_SEL = test_item.get_32bit_reg_address("RW_TEST0_SEL", df_register)
        RW_TEST0_EN = test_item.get_32bit_reg_address("RW_TEST0_EN", df_register)
        RW_OSC_24M_TRIM = test_item.get_32bit_reg_address("RW_OSC_24M_TRIM", df_register)
    except Exception as e:
        print(f"❌ 寄存器地址获取失败: {e}")
        return
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")

    print("请将示波器的CH1探头接入P13引脚, 按下enter键继续...")
    input()
    time.sleep(1)
    try:
        #初始化、写入测试指令
        instr.swd_modify_register(Pad1_ana_en, 0x1, 11, 11)
        instr.swd_modify_register(RW_TEST0_SEL, 0x06, 16, 18)
        instr.swd_modify_register(RW_TEST0_EN, 0x1, 15, 15)
        time.sleep(0.2)
    except Exception as e:
        print(f"❌ 寄存器写入/万用表读取失败: {e}")
        return
    
    typ_val_period = 1000.0 / float(typ_val)   #测试项写的是频率, 需要转换成周期
    print(f"测试目标频率为type_val={typ_val * 1000} Hz, 目标周期为{typ_val_period:.3f} us")

    instr.osc_auto_set()
    time.sleep(2)
    #准备测试数据
    instr.swd_modify_register(RW_OSC_24M_TRIM, 0x20, 8, 13)
    time.sleep(0.5)
    OSC24MA_freq, OSC24MA_period, OSC24MA_duty = instr.osc_measure_frequency("CH1")
    print(f"OSC24M频率A: {OSC24MA_freq:.3f} Hz, 周期: {OSC24MA_period:.3f} us")

    instr.swd_modify_register(RW_OSC_24M_TRIM, 0x21, 8, 13)
    time.sleep(0.5)
    OSC24MB_freq, OSC24MB_period, OSC24MB_duty = instr.osc_measure_frequency("CH1")
    print(f"OSC24M频率B: {OSC24MB_freq:.3f} Hz, 周期: {OSC24MB_period:.3f} us")

    delta_osc = OSC24MB_period - OSC24MA_period
    min_error = float('inf')

    if abs(delta_osc) < 1e-6:
        print(f"⚠️ OSC24M修调码线性插值失败(分母为0), 切换到全局遍历模式(0~63)")
        search_range = range(30, 38)
    else:
        TRIMCODE = (typ_val_period - OSC24MA_period) / delta_osc + 32.0
        TRIM_CALC = round(TRIMCODE)
        search_range = range(TRIM_CALC - 4, TRIM_CALC + 6)
        print(f"OSC24M修调码计算结果: {TRIMCODE:.2f}, 四舍五入取整: {TRIM_CALC}, 搜索范围: {search_range.start} ~ {search_range.stop - 1}")

    for test_trim in search_range:
        if test_trim < 0 or test_trim > 63:
            continue
        instr.swd_modify_register(RW_OSC_24M_TRIM, test_trim, 8, 13)
        time.sleep(0.5)
        #复测，计算test_period与typ_val的绝对误差，取最小值对应的最优trim
        test_freq, test_period, test_duty = instr.osc_measure_frequency("CH1")
        test_error = abs(test_period - typ_val_period)
        print(f"测试修调码 {test_trim}: 频率={test_freq:.3f} Hz, 周期={test_period:.3f} us, 与目标周期误差={test_error:.3f} us")

        if test_error < min_error:
            min_error = test_error
            best_trim = test_trim

    instr.swd_modify_register(RW_OSC_24M_TRIM, best_trim, 8, 13)
    time.sleep(0.5)
    global_use["HIRC_TRIM"] = RW_OSC_24M_TRIM
    final_freq, final_period, final_duty = instr.osc_measure_frequency("CH1")
    print(f"最终OSC24M频率: {final_freq:.3f} Hz, 周期: {final_period:.3f} us")
    print(f"24MHZ_Value:{instr.swd_read_register(RW_OSC_24M_TRIM)}")

    final_freq = final_freq / 1000      #注意示波器测试出来是的频率单位是Hz, 需要转换到KHz进行比较
    if final_freq >= min_val and final_freq <= max_val:
        print("OSC24M Test pass")
    else:
        print("OSC24M Test failed")
        print(f"修调范围为: {min_val}~{max_val} {unit}, 实际测量值: {final_freq:.3f} {unit}")
    """保存结果"""

def I_NORMAL_SLEEP_Test():
    """5.5 Normal电流 自动化修调测试"""
    test_item_name = "INORMAL"
    row = df_test_data[df_test_data["测试项"] == test_item_name]
    if row.empty:
        print(f"未找到测试项: {test_item_name}")
        return False
    
    try:
        typ_val = float(row["Typ."].values[0])
        unit = str(row["Unit"].values[0])
    except (IndexError, ValueError) as e:
        print(f"❌ VBG测试参数读取失败: {e}")
        return False
    
    # 获取寄存器地址
    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register) 
        RW_ANA_EN = test_item.get_32bit_reg_address("RW_ANA_EN", df_register) 
        RW_HFCK_EN = test_item.get_32bit_reg_address("RW_HFCK_EN", df_register) 
    except Exception as e:
        print(f"❌ VBG寄存器地址获取失败: {e}")
        return False

    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    time.sleep(1)

    instr.swd_modify_register(Pad1_ana_en, 0xffffff, 8, 31)

    """这里需要循环将每个引脚接入万用表, 挨个测量每个引脚的工作电流"""
    """保存结果"""

    instr.swd_modify_register(RW_ANA_EN, 0x0, 8, 8)
    instr.swd_modify_register(RW_HFCK_EN, 0x0, 8, 8)
    """这里需要循环将每个引脚接入万用表, 挨个测量每个引脚的睡眠电流"""
    """保存结果"""

def VBG_Test():
    """5.6 VBG 自动化修调测试"""
    test_item_name = "VBG"
    # 读取VBG测试规格参数
    row = df_test_data[df_test_data["测试项"] == test_item_name]
    if row.empty:
        print(f"未找到测试项: {test_item_name}")
        return False
    try:
        min_val = float(row["Min."].values[0])  # VBG最小值
        typ_val = float(row["Typ."].values[0])  # VBG典型值( 目标值) 
        max_val = float(row["Max."].values[0])  # VBG最大值
        unit = str(row["Unit"].values[0])       # 电压单位( 如V/mV) 
    except (IndexError, ValueError) as e:
        print(f"❌ VBG测试参数读取失败: {e}")
        return False
    
    # 获取VBG相关寄存器地址( 从register.xlsx解析) 
    try:
        Pad2_ana_en = test_item.get_32bit_reg_address("Pad2_ana_en", df_register) 
        RW_TEST1_SEL = test_item.get_32bit_reg_address("RW_TEST1_SEL", df_register)
        RW_TEST1_EN = test_item.get_32bit_reg_address("RW_TEST1_EN", df_register) 
        RW_VBG_TRIM = test_item.get_32bit_reg_address("RW_VBG_TRIM", df_register)
    except Exception as e:
        print(f"❌ VBG寄存器地址获取失败: {e}")
        return False
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    print("将ADC模块输出线接入VCC_S引脚, 万用表表笔接入P20引脚, 按下enter键继续...")
    input()
    time.sleep(1)
    
    try:
        # 写入使能/模式配置
        instr.swd_modify_register(Pad2_ana_en, 0xff, 16 ,23)
        instr.swd_modify_register(RW_TEST1_SEL, 0x0, 19, 22)

        # 这里VCC_S输入 1200mV电压
        instr.serial_send("spi_cmd_set_val mV 1200\r\n")
        time.sleep(1)
        
        instr.swd_modify_register(RW_TEST1_EN, 0x1, 16, 16) 
        time.sleep(0.2)  # 等待模块稳定
    except Exception as e:
        print(f"❌ VBG寄存器初始化失败: {e}")
        return False
    
    VS = instr.dmm_measure_single_voltage()
    voffset = VS - 1200.0

    # 这里需要检测引脚输出
    instr.swd_modify_register(RW_TEST1_SEL, 0x6, 19, 22)
    time.sleep(0.2)
    
    best_trim = 0
    min_error = 9999.0
    final_vbg = 0.0

    for trim in range(0, 8):  # 3位TRIM: 000~111
        # 写入TRIM值
        instr.swd_modify_register(RW_VBG_TRIM, trim, 24, 26)
        time.sleep(0.1)
        # 测量+计算真实电压
        v_meas = instr.dmm_measure_single_voltage()
        vbg_real = v_meas - voffset
        error = abs(vbg_real - typ_val)

        print(f"TRIM={trim:>2} | 测量={v_meas:>6.2f}mV | 真实={vbg_real:>6.2f}mV | 误差={error:>5.2f}mV")

        # 筛选: 符合规格 + 误差最小
        if min_val <= vbg_real <= max_val and error < min_error:
            min_error = error
            best_trim = trim
            final_vbg = vbg_real

    instr.swd_modify_register(RW_VBG_TRIM, best_trim, 24, 26)
    print(f"\n🎯 VBG 修调完成! 最优TRIM={best_trim} | 最终电压={final_vbg:.2f}mV")
    if min_val <= final_vbg <= max_val:
        print("✅ VBG 修调结果: PASS")
        return True
    else:
        print("❌ VBG 修调结果: FAIL")
        return False
    """保存结果"""

def VPTAT_Test():
    """5.6 VPTAT 自动化修调测试"""
    test_item_name = "VPTAT"
    # 读取VPTAT测试规格参数
    row = df_test_data[df_test_data["测试项"] == test_item_name]
    if row.empty:
        print(f"未找到测试项: {test_item_name}")
        return False
    try:
        typ_val = float(row["Typ."].values[0])  # 典型值( 目标值) 
        unit = str(row["Unit"].values[0])       # 电压单位( 如V/mV) 
    except (IndexError, ValueError) as e:
        print(f"❌ VPTAT测试参数读取失败: {e}")
        return False
    
    # 获取VBG相关寄存器地址( 从register.xlsx解析) 
    try:
        Pad2_ana_en = test_item.get_32bit_reg_address("Pad2_ana_en", df_register) 
        RW_TEST1_SEL = test_item.get_32bit_reg_address("RW_TEST1_SEL", df_register)
        RW_TEST1_EN = test_item.get_32bit_reg_address("RW_TEST1_EN", df_register) 
        RW_VPTAT_TRIM = test_item.get_32bit_reg_address("RW_VPTAT_TRIM", df_register)
    except Exception as e:
        print(f"❌ VPTAT寄存器地址获取失败: {e}")
        return False
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    print("将ADC模块输出线接入VCC_S引脚, 万用表表笔接入P20引脚, 按下enter键继续...")
    input()
    time.sleep(1)

    try:
        # 写入使能/模式配置
        instr.swd_modify_register(Pad2_ana_en, 0xff, 16 ,23)
        instr.swd_modify_register(RW_TEST1_SEL, 0x0, 19, 22)

        # 这里VCC_S输入 700mV电压
        instr.serial_send("spi_cmd_set_val mV 700\r\n")
        time.sleep(1)

        instr.swd_modify_register(RW_TEST1_EN, 0x1, 16, 16) 
        time.sleep(0.2)  # 等待模块稳定
    except Exception as e:
        print(f"❌ VBG寄存器初始化失败: {e}")
        return False
    
    VS = instr.dmm_measure_single_voltage()
    voffset = VS - 700.0

    # 检测引脚输出
    instr.swd_modify_register(RW_TEST1_SEL, 0x4, 19, 22)
    time.sleep(0.2)
    
    best_trim = 0
    min_error = 9999.0
    final_vptat = 0.0
    
    for trim in range(0, 4):  # 2位TRIM: 00~11
        # 写入TRIM值
        instr.swd_modify_register(RW_VPTAT_TRIM, trim, 28, 29)
        time.sleep(0.1)
        # 测量+计算真实电压
        v_meas = instr.dmm_measure_single_voltage()
        vptat_real = v_meas - voffset
        error = abs(vptat_real - typ_val)

        print(f"TRIM={trim:>2} | 测量={v_meas:>6.2f}mV | 真实={vptat_real:>6.2f}mV | 误差={error:>5.2f}mV")

        # 筛选: 符合规格 + 误差最小
        if vptat_real >= typ_val and error < min_error:
            min_error = error
            best_trim = trim
            final_vptat = vptat_real

    instr.swd_modify_register(RW_VPTAT_TRIM, best_trim, 28, 29)
    print(f"\n🎯 VPTAT 修调完成! 最优TRIM={best_trim} | 最终电压={final_vptat:.2f}mV")
    if final_vptat >= typ_val:
        print("✅ VPTAT 修调结果: PASS")
        return True
    else:
        print("❌ VPTAT 修调结果: FAIL")
        return False
    """保存结果"""

def VREF_Test():
    """5.7 VREF 自动化修调测试"""
    test_item_name = "VREF"
    #获取测试数据、寄存器地址
    row = df_test_data[df_test_data["测试项"] == test_item_name]
    if row.empty:
        print(f"未找到测试项: {test_item_name}")
        return
    try:
        min_val = float(row["Min."].values[0])
        typ_val = float(row["Typ."].values[0])
        max_val = float(row["Max."].values[0])
        unit = str(row["Unit"].values[0])
    except (IndexError, ValueError) as e:
        print(f"❌ 测试参数读取失败: {e}")
        return

    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register)
        RW_VREF_EN = test_item.get_32bit_reg_address("RW_VREF_EN", df_register)
        RW_VHALF_EN = test_item.get_32bit_reg_address("RW_VHALF_EN", df_register)
        RW_VREF_SEL = test_item.get_32bit_reg_address("RW_VREF_SEL", df_register)
        RW_VREF_TRIM = test_item.get_32bit_reg_address("RW_VREF_TRIM", df_register)
        RW_VBG_TRIM = test_item.get_32bit_reg_address("RW_VBG_TRIM", df_register)
    except Exception as e:
        print(f"❌ 寄存器地址获取失败: {e}")
        return

    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    print("请将万用表表笔接入P11引脚, 按下enter键继续...")
    input()
    time.sleep(1)

    try:
        #初始化、写入测试指令
        instr.swd_modify_register(Pad1_ana_en, 0xff, 8, 15)
        instr.swd_modify_register(RW_VREF_EN, 0x1, 8, 8)
        instr.swd_modify_register(RW_VHALF_EN, 0x1, 7, 7)
        instr.swd_modify_register(RW_VREF_SEL, 0x3, 0, 1)
        instr.swd_modify_register(RW_VBG_TRIM, 0x7, 24, 26)
        time.sleep(0.2)

        # 准备所需数据
        instr.swd_modify_register(RW_VREF_TRIM, 0x0, 0, 5)
        
        time.sleep(0.2)
        VREFA = instr.dmm_measure_single_voltage()
        print(f"初始VREFA测量值: {VREFA:.2f}{unit}")

        instr.swd_modify_register(RW_VREF_TRIM, 0x10, 0, 5)
        time.sleep(0.2)
        VREFB = instr.dmm_measure_single_voltage()
        print(f"初始VREFB测量值: {VREFB:.2f}{unit}")
    except Exception as e:
        print(f"❌ 寄存器写入/万用表读取失败: {e}")
        return

    delta_v = (VREFB - VREFA) / 16.0  
    if abs(delta_v) < 1e-6:  # 避免除以0
        print(f"⚠️  delta_v={delta_v}{unit}, 修调码计算失败, 使用默认值32")
        best_trim = 32
    else:
        best_trim = int(round((typ_val - VREFA) / delta_v + 0))
        #best_trim = max(0, min(63, raw_trim))  # 强制限定在0~63( 6位TRIM最大值) 


    min_error = float('inf')  # 初始化无穷大, 避免魔法数999
    final_trim = best_trim
    final_voltage = 0

    """"""
    for i in range(-5, 5):
        try:
            trim_candidate = best_trim + i
            instr.swd_modify_register(RW_VREF_TRIM, trim_candidate, 0, 5)
            time.sleep(0.2)
            current_v = instr.dmm_measure_single_voltage()
            print(f"测试TRIM={trim_candidate}: 测量电压={current_v:.2f}{unit}")
            current_error = abs(current_v - typ_val)

            # 如果这个值误差更小, 就记录下来
            if current_error < min_error:
                min_error = current_error
                final_trim = trim_candidate
                final_voltage = current_v

        except Exception as e:
            print(f"⚠️  TRIM={trim_candidate} 测试失败: {e}, 跳过该值")
            continue

    # 将最优调吗写入寄存器
    instr.swd_modify_register(RW_VREF_TRIM, final_trim, 0, 5)
    global_use["VREF_TRIM"] = RW_VREF_TRIM
    global_use["Voffset_P"] = instr.dmm_measure_single_voltage()  # 存储Voffset_P供ADC测试使用
    print(f"\n🎯 VREF 修调完成! 最优TRIM={final_trim} | 最终电压={final_voltage:.2f}{unit}")
    if min_val <= final_voltage <= max_val:
        print("✅ 测试结果: PASS")
        return True  # 返回测试结果, 方便上层调用
    else:
        print(f"❌ 测试结果: FAIL(范围要求: {min_val}~{max_val}{unit})")
        return False
    """保存结果"""

def DAC0_DAC1_DNL_Test():
    """5.8 DAC0,DAC1 自动化修调测试"""
    dac0_test_item_name = "DAC0_DNL"
    dac1_test_item_name = "DAC1_DNL"
    dac0_row = df_test_data[df_test_data["测试项"] == dac0_test_item_name]
    dac1_row = df_test_data[df_test_data["测试项"] == dac1_test_item_name]
    if dac0_row.empty and dac1_row.empty:
        print(f"未找到测试项: {dac0_test_item_name} and {dac1_test_item_name}")
        return
    try:
        dac0_min_val = float(dac0_row["Min."].values[0])
        dac0_max_val = float(dac0_row["Max."].values[0])

        dac1_min_val = float(dac1_row["Min."].values[0])
        dac1_max_val = float(dac1_row["Max."].values[0])

        unit = str(dac0_row["Unit"].values[0])
    except (IndexError, ValueError) as e:
        print(f"❌ 测试参数读取失败: {e}")
        return
    
    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register)
        Pad2_ana_en = test_item.get_32bit_reg_address("Pad2_ana_en", df_register)
        RW_TEST1_SEL = test_item.get_32bit_reg_address("RW_TEST1_SEL", df_register)
        RW_TEST1_EN = test_item.get_32bit_reg_address("RW_TEST1_EN", df_register)
        RW_VREF_EN = test_item.get_32bit_reg_address("RW_VREF_EN", df_register)
        RW_VREF_SEL = test_item.get_32bit_reg_address("RW_VREF_SEL", df_register)
        RW_VREF_TRIM = test_item.get_32bit_reg_address("RW_VREF_TRIM", df_register)
        RW_DAC0_DIN = test_item.get_32bit_reg_address("RW_DAC0_DIN", df_register)
        RW_DAC1_DIN = test_item.get_32bit_reg_address("RW_DAC1_DIN", df_register)
        RW_DAC_EN = test_item.get_32bit_reg_address("RW_DAC_EN", df_register)
    except Exception as e:
        print(f"❌ DAC寄存器地址获取失败: {e}")
        return False
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    print("请将ADC模块输出线接入VCC_S引脚, 万用表表笔接入P20引脚, 按下enter键继续...")
    input()
    time.sleep(1)

    #初始化配置
    # 1. 相关引脚和VREF 配置
    instr.swd_modify_register(Pad1_ana_en, 0xff, 8, 15)
    instr.swd_modify_register(Pad2_ana_en, 0xff, 16, 23)
    instr.swd_modify_register(RW_VREF_EN, 0x01, 8, 8)
    instr.swd_modify_register(RW_VREF_SEL, 0x03, 0, 1)
    instr.swd_modify_register(RW_VREF_TRIM, global_use["VREF_TRIM"], 0, 5)

    instr.swd_modify_register(RW_TEST1_SEL, 0x0, 19, 22)

    """这里需要提示在VCC_S引脚接入2000mV基准电压"""
    instr.serial_send("spi_cmd_set_val mV 2000\r\n")
    print("⚠️ 请在VCC_S引脚接入2000mV基准电压后, 按回车继续...")
    input()

    instr.swd_modify_register(RW_TEST1_EN, 0x01, 16, 16)
    time.sleep(0.2)
    VS = instr.dmm_measure_single_voltage()
    Voffset_M = VS - 2000.0
    global_use["Voffset_M"] = Voffset_M     #存入Voffset_M让下一个函数使用

    instr.swd_modify_register(RW_TEST1_SEL, 0x04, 19, 22)
    time.sleep(0.1)
    p20_0100_v = instr.dmm_measure_single_voltage()
    print(f"📊 SEL=0100B时, P20输出 = {p20_0100_v} mV")

    instr.swd_modify_register(RW_TEST1_SEL, 0x03, 19, 22)
    time.sleep(0.1)
    p20_0011_v = instr.dmm_measure_single_voltage()
    print(f"📊 SEL=0011B时, P20输出 = {p20_0011_v} mV")

    dac0_code_list = [1, 2, 4, 8, 16, 32, 64, 128, 256]
    dac0_voltage_list = []  # 存储实测电压Vreal
    dac0_dnl_list = []      # 存储【DNL值】

    instr.swd_modify_register(RW_DAC_EN, 0x01, 2, 2)
    # 2. 循环测试每个码值
    for code in dac0_code_list:
        instr.swd_modify_register(RW_DAC0_DIN, code, 0, 8)
        time.sleep(0.1)

        # 读取实测电压
        Vreal0 = instr.dmm_measure_single_voltage()
        dac0_voltage_list.append(Vreal0)  # 保存电压

        # 计算理论电压、误差、DNL
        Videal0 = code / 512 * 4000
        dnl = Videal0 - Vreal0 - Voffset_M
        dac0_dnl_list.append(dnl)

    # 4. 计算最终DNL( 取最大值) 
    DAC0_DNL = max(abs(dnl) for dnl in dac0_dnl_list)
    #DAC0_dnl_min, DAC0_dnl_max = min(dac0_dnl_list), max(dac0_dnl_list)
    if dac0_min_val <= DAC0_DNL and dac0_max_val >= DAC0_DNL:
        print(f"\n✅ DAC0 DNL验证通过! 最大DNL={DAC0_DNL:.3f} LSB(阈值±2 LSB)")
    else:
        print(f"\n❌ DAC0 DNL验证失败! 最大DNL={DAC0_DNL:.3f} LSB(超出±2 LSB)")

    instr.swd_modify_register(RW_TEST1_SEL, 0x02, 19, 22)
    time.sleep(0.2)
    p20_0010_v = instr.dmm_measure_single_voltage()
    print(f"📊 SEL=0010B时, P20输出 = {p20_0010_v} mV")
    Voffset = Voffset_M     #这个 Voffset 需修改

    dac1_code_list = [1, 2, 4, 8, 16, 32]
    dac1_voltage_list = []  # 存储【实测电压Vreal】
    dac1_dnl_list = []      # 存储【DNL值】

    # 2. 循环测试每个码值
    for code in dac1_code_list:
        instr.swd_modify_register(RW_DAC1_DIN, code, 16, 21)
        time.sleep(0.1)

        # 读取实测电压
        Vreal = instr.dmm_measure_single_voltage()
        dac1_voltage_list.append(Vreal)  # 保存电压

        # 计算理论电压、误差、DNL
        Videal1 = code / 64 * 4000
        dnl = Videal1 - Vreal - Voffset
        dac1_dnl_list.append(dnl)

    # 4. 计算最终DNL( 取最大值) 
    DAC1_DNL = max(abs(dnl) for dnl in dac1_dnl_list)
    #DAC1_dnl_min, DAC1_dnl_max = min(dac1_dnl_list), max(dac1_dnl_list)
    if dac1_min_val <= DAC1_DNL and dac1_max_val >= DAC1_DNL:
        print(f"\n✅ DAC1 DNL验证通过! 最大DNL={DAC1_DNL:.3f} LSB(阈值±2 LSB)")
    else:
        print(f"\n❌ DAC1 DNL验证失败! 最大DNL={DAC1_DNL:.3f} LSB(超出±2 LSB)")
    """保存结果"""

def ADC_DNL_Test():
    """5.9 ADC 测试"""
    adc_test_item_name = "ADC_DNL"
    row = df_test_data[df_test_data["测试项"] == adc_test_item_name]
    if row.empty:
        print("未找到测试项: {adc_test_item_name}")
        return
    try:
        adc_min_val = float(row["Min."].values[0])
        adc_max_val = float(row["Max."].values[0])
        unit = str(row["Unit"].values[0])
    except (IndexError, ValueError) as e:
        print(f"❌ 测试参数读取失败: {e}")
        return
    
    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register)
        Pad2_ana_en = test_item.get_32bit_reg_address("Pad2_ana_en", df_register)
        RW_ADC_CLK = test_item.get_32bit_reg_address("RW_ADC_CLK", df_register)
        RW_ADC_CTRL_EN = test_item.get_32bit_reg_address("RW_ADC_CTRL_EN", df_register)
        RW_VREF_EN = test_item.get_32bit_reg_address("RW_VREF_EN", df_register)
        # RW_VREF_SEL = test_item.get_32bit_reg_address("RW_VREF_SEL", df_register)
        # RW_VREF_TRIM = test_item.get_32bit_reg_address("RW_VREF_TRIM", df_register)
        RW_ADC_EN = test_item.get_32bit_reg_address("RW_ADC_EN", df_register)
        cfg_I_sampl_en = test_item.get_32bit_reg_address("cfg_I_sampl_en", df_register)
        CH6EN = test_item.get_32bit_reg_address("CH6EN", df_register)
        ADCBSY = test_item.get_32bit_reg_address("ADCBSY", df_register)
        ADC_SCYC = test_item.get_32bit_reg_address("ADC_SCYC", df_register)
        ADC6_DR = test_item.get_32bit_reg_address("ADC6_DR", df_register)
    except Exception as e:
        print(f"❌ ADC寄存器地址获取失败: {e}")
        return False
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    time.sleep(1)
    
    instr.swd_modify_register(Pad1_ana_en, 0xff, 8, 15)
    instr.swd_modify_register(Pad2_ana_en, 0xff, 16, 23)
    instr.swd_modify_register(0x40001424, 0xffffffff, 0, 31)
    instr.swd_modify_register(RW_ADC_CLK, 0x1, 20, 20)
    # instr.swd_modify_register(RW_VREF_SEL, 0x03, 0, 1)
    instr.swd_modify_register(RW_VREF_EN, 0x01, 8, 8)
    # instr.swd_modify_register(RW_VREF_TRIM, global_use["VREF_TRIM"], 0, 5)
    instr.swd_modify_register(RW_ADC_EN, 0x01, 1, 1)
    instr.swd_modify_register(RW_ADC_CTRL_EN, 0x01, 7, 7)
    instr.swd_modify_register(cfg_I_sampl_en, 0x0, 14, 14)
    instr.swd_modify_register(CH6EN, 0x1, 6, 6)
    instr.swd_modify_register(ADC_SCYC, 0x7, 10, 13)
    
    time.sleep(0.1)

    print("⚠️ 请在P20引脚接入DAC模块输出线后, 按回车继续...")

    input()

    Guding = 1000

    # 选择测试电压模式
    print("\n请选择ADC DNL测试电压模式:")
    print("1. 固定测试电压 (1000mV)")
    print("2. 随机测试电压 (10组0.1~4.0V)")
    print("3. 线性扫描测试电压 (500mV~4000mV, 步进100mV)")
    choice = input("请输入选项 (1/2/3): ").strip()

    random.seed(42)  # 固定随机种子
    if choice == '1':
        vin_list = [1.0] * 10
        print(f"已选择固定测试电压模式: 1000mV, 共{len(vin_list)}个测试点")
    elif choice == '2':
        vin_list = [round(random.uniform(0.5, 4.0), 3) for _ in range(10)]
        print(f"已选择随机测试电压模式: 共{len(vin_list)}个测试点")
        for i, v in enumerate(vin_list):
            print(f"  点{i+1}: {v*1000:.0f}mV")
    else:
        # 线性扫描模式: 让用户设置步进值和记录点数
        step_input = input("请输入线性扫描步进值(mV, 默认100mV): ").strip()
        try:
            step_mv = float(step_input) if step_input else 100.0
        except ValueError:
            print(f"输入无效: {step_input}, 使用默认步进100mV")
            step_mv = 100.0

        points_input = input("请输入记录点数(默认36个点, 500mV~4000mV): ").strip()
        try:
            num_points = int(points_input) if points_input else 36
        except ValueError:
            print(f"输入无效: {points_input}, 使用默认36个点")
            num_points = 36

        step_v = step_mv / 1000.0  # 转换为V单位
        vin_list = []
        for i in range(num_points):
            vin = 0.5 + i * step_v
            if vin > 4.0:
                break
            vin_list.append(round(vin, 3))
        print(f"已选择线性扫描测试电压模式: 起始500mV, 步进{step_mv:.0f}mV, 记录{len(vin_list)}个点")

    ADC_LSB_list = []
    dnl_errors = []
    raw_data_list = []  # 保存每次测试的原始数据

    # 预计算: 将不随循环变化的值提到外面
    adc_lsb_value = global_use["Voffset_P"] / 4095  # 每个LSB对应的mV值
    voffset_p = global_use["Voffset_P"]  # Voffset_P引用

    for idx, vin in enumerate(vin_list):

        vin_set = vin * 1000  # 转换为mV单位
        instr.serial_send(f"spi_cmd_set_val mV {int(vin_set)}\r\n")
        # DAC稳定等待时间(10ms, DAC响应通常在ms级)
        time.sleep(0.01)

        # 首次测量用完整配置，后续用快速测量(跳过CONFigure重复配置,省~200ms/次)
        if idx == 0:
            vol = instr.dmm_measure_single_voltage()
        else:
            vol = instr.dmm_quick_measure_voltage()

        # 启动ADC转换(无需等待，后续立即读取结果)
        instr.swd_modify_register(ADCBSY, 0x01, 6, 6)

        # ADC结果已就绪，直接读取
        get_adc_val = instr.swd_read_register(ADC6_DR) & 0xfff

        #计算实际输出电压
        vreal = get_adc_val / 4095 * voffset_p
       
        #计算DNL误差(使用预计算的adc_lsb_value)
        dnl_error = vol - vreal
        dnl_error_lsb = dnl_error / adc_lsb_value
        dnl_errors.append(dnl_error)
        ADC_LSB_list.append(dnl_error_lsb)
        raw_data_list.append({
            "vin_mV": vin_set,
            "adc_raw": get_adc_val,
            "vreal_mV": vreal,
            "vol_mV": vol,
            "dnl_lsb": dnl_error_lsb
        })
        print(f" ADC原始数据: {get_adc_val} 转换电压: {vreal:.3f} mV | 实测电压: {vol:.2f} mV | DNL误差: {dnl_error_lsb:.3f}")

    adc_dnl_max = max(ADC_LSB_list)  # 取最大LSB误差作为最终DNL指标
    print(f"\n📊 ADC DNL最大误差: {adc_dnl_max:.3f}")
    print(f"⚠️ ADC DNL测试结果要求在 {adc_min_val} ~ {adc_max_val} 之间")
    if adc_min_val <= adc_dnl_max <= adc_max_val:
        print(f"✅ ADC_DNL测试通过! ")
    else:
        print(f"❌ ADC_DNL测试失败! 超出规格范围! ")

    # 保存结果到Excel
    try:
        import openpyxl
        from openpyxl import Workbook

        # 提示输入文件名
        default_name = "ADC_DNL_Test_Result"
        file_name = input(f"\n请输入保存文件名(不含扩展名, 默认{default_name}): ").strip()
        if not file_name:
            file_name = default_name

        # 获取桌面路径
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop, f"{file_name}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "ADC_DNL_Test_Data"

        # 写入表头
        headers = ["测试电压(mV)", "ADC原始数据", "转换电压(mV)", "实测电压(mV)", "DNL误差(LSB)"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # 写入详细测试数据
        for row_idx, data in enumerate(raw_data_list, 2):
            ws.cell(row=row_idx, column=1, value=data["vin_mV"])
            ws.cell(row=row_idx, column=2, value=data["adc_raw"])
            ws.cell(row=row_idx, column=3, value=round(data["vreal_mV"], 3))
            ws.cell(row=row_idx, column=4, value=round(data["vol_mV"], 2))
            ws.cell(row=row_idx, column=5, value=round(data["dnl_lsb"], 3))

        # 写入汇总信息
        summary_row = len(raw_data_list) + 3
        ws.cell(row=summary_row, column=1, value="汇总信息")
        ws.cell(row=summary_row + 1, column=1, value="ADC DNL最大误差(LSB)")
        ws.cell(row=summary_row + 1, column=2, value=round(adc_dnl_max, 3))
        ws.cell(row=summary_row + 2, column=1, value="测试点数")
        ws.cell(row=summary_row + 2, column=2, value=len(raw_data_list))
        ws.cell(row=summary_row + 3, column=1, value="规格下限")
        ws.cell(row=summary_row + 3, column=2, value=adc_min_val)
        ws.cell(row=summary_row + 4, column=1, value="规格上限")
        ws.cell(row=summary_row + 4, column=2, value=adc_max_val)
        ws.cell(row=summary_row + 5, column=1, value="测试结果")
        ws.cell(row=summary_row + 5, column=2, value="PASS" if adc_min_val <= adc_dnl_max <= adc_max_val else "FAIL")

        wb.save(file_path)
        print(f"\n✅ 测试结果已保存至桌面: {file_path}")
    except ImportError:
        print("\n⚠️ 未安装openpyxl模块, 无法保存Excel文件, 请执行: pip install openpyxl")
    except Exception as e:
        print(f"\n❌ 保存Excel文件失败: {e}")
    """保存结果"""

    # 测试完成, 关闭通道和ADC
    instr.swd_write_register(CH6EN, 0x00 << 6)
    instr.swd_write_register(RW_ADC_EN, 0x00 << 1)

def Voffset_VS_ERROR_Test():
    """5.10 VCC_S 自动化脚本测试"""
    # 定义测试项名称
    voffset_test_item = "Voffset"
    vs_error_test_item = "VS_ERROR"
    
    # 从测试表格读取规格参数
    row_voffset = df_test_data[df_test_data["测试项"] == voffset_test_item]
    row_vs_error = df_test_data[df_test_data["测试项"] == vs_error_test_item]
    
    if row_voffset.empty or row_vs_error.empty:
        print(f"未找到测试项: {voffset_test_item} 或 {vs_error_test_item}")
        return
    
    try:
        # 读取Voffset_VS规格
        voffset_min = float(row_voffset["Min."].values[0])
        voffset_max = float(row_voffset["Max."].values[0])
        voffset_unit = str(row_voffset["Unit"].values[0])
        # 读取VS_ERROR规格
        vs_error_min = float(row_vs_error["Min."].values[0])
        vs_error_max = float(row_vs_error["Max."].values[0])
        vs_error_unit = str(row_vs_error["Unit"].values[0])
    except (IndexError, ValueError) as e:
        print(f"❌ VCC_S测试参数读取失败: {e}")
        return
    
    # 获取寄存器地址( 与项目统一寄存器名一致) 
    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register)
        Pad2_ana_en = test_item.get_32bit_reg_address("Pad2_ana_en", df_register)
        RW_TEST1_SEL = test_item.get_32bit_reg_address("RW_TEST1_SEL", df_register)
        RW_ADC_CLK = test_item.get_32bit_reg_address("RW_ADC_CLK", df_register)
        RW_ADC_CTRL_EN = test_item.get_32bit_reg_address("RW_ADC_CTRL_EN", df_register)
        RW_TEST1_EN = test_item.get_32bit_reg_address("RW_TEST1_EN", df_register)
        RW_VREF_EN = test_item.get_32bit_reg_address("RW_VREF_EN", df_register)
        RW_VREF_SEL = test_item.get_32bit_reg_address("RW_VREF_SEL", df_register)
        RW_VREF_TRIM = test_item.get_32bit_reg_address("RW_VREF_TRIM", df_register)
        cfg_I_sampl_en = test_item.get_32bit_reg_address("cfg_I_sampl_en", df_register)
        RW_ADC_EN = test_item.get_32bit_reg_address("RW_ADC_EN", df_register)
        CH10EN = test_item.get_32bit_reg_address("CH10EN", df_register)
        ADCBSY = test_item.get_32bit_reg_address("ADCBSY", df_register)
        ADC_SCYC = test_item.get_32bit_reg_address("ADC_SCYC", df_register)
        ADC10_DR = test_item.get_32bit_reg_address("ADC10_DR", df_register)
    except Exception as e:
        print(f"❌ VCC_S寄存器地址获取失败: {e}")
        return False
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    print("请将DAC模块输出线接入VCC_S引脚, 万用表表笔接入P20引脚, 按下enter键继续...")
    input()
    time.sleep(1)
    
    instr.swd_modify_register(Pad1_ana_en, 0xff, 8, 15)
    instr.swd_modify_register(Pad2_ana_en, 0xff, 16, 23)
    instr.swd_modify_register(RW_VREF_SEL, 0x03, 0, 1)
    instr.swd_modify_register(RW_VREF_EN, 0x01, 8, 8)
    instr.swd_modify_register(RW_VREF_TRIM, global_use["VREF_TRIM"], 0, 5)
    instr.swd_modify_register(RW_TEST1_SEL, 0x00, 19, 22)
    time.sleep(0.1)

    instr.serial_send("spi_cmd_set_val mV 3400\r\n")
    print("⚠️ 请在VCC_S引脚接入3400mV电压后, 按回车继续...")
    input()

    instr.swd_modify_register(RW_TEST1_EN, 0x01, 16, 16)
    time.sleep(0.1)
    VS = instr.dmm_measure_single_voltage()
    Voffset= VS - 3400

    if(Voffset >= voffset_min and Voffset <= voffset_max):
        print(f"✅ Voffset 测试通过")
    else:
        print(f"❌ Voffset 测试失败(规格: {voffset_min}~{voffset_max}{voffset_unit})")

    #开始监测VS电压
    instr.swd_modify_register(RW_TEST1_EN, 0x00, 16, 16)
    instr.swd_modify_register(cfg_I_sampl_en, 0x0, 14, 14)
    instr.swd_modify_register(RW_ADC_CLK, 0x01, 20, 20)
    instr.swd_modify_register(RW_ADC_CTRL_EN, 0x01, 7, 7)
    instr.swd_modify_register(RW_ADC_EN, 0x01, 1, 1)
    instr.swd_modify_register(CH10EN, 0x01, 10, 10)
    time.sleep(0.1)

    instr.swd_modify_register(ADCBSY, 0x01, 6, 6)
    time.sleep(0.1)
    get_adc_val = instr.swd_read_register(ADC10_DR)

    CODEreal = get_adc_val & 0xFFF
    CODEideal = VS / 4000 * 4096
    VS_ERROR = CODEreal - CODEideal
    if VS_ERROR >= vs_error_min and VS_ERROR <= vs_error_max:
        print(f"✅ VS_ERROR 测试通过")
    else:
        print(f"❌ VS_ERROR 测试失败(规格: {vs_error_min}~{vs_error_max}{vs_error_unit})")
    """保存结果"""

    instr.swd_modify_register(RW_ADC_EN, 0x00, 1, 1)
    instr.swd_modify_register(CH10EN, 0x00, 10, 10)
    instr.swd_modify_register(RW_VREF_EN, 0x0, 8, 8)
    instr.swd_modify_register(RW_TEST1_EN, 0x0, 16, 16)

def PGA_ADC_SH_PGA_Test():
    """5.11 PGA+ADC& SH+PGA测试"""
    # 定义测试项名称
    VAZ0_test_item = "VAZ0"
    VAZ1_test_item = "VAZ1"
    VAZ2_test_item = "VAZ2"
    IUerror_test_item = "IUerror"
    VUerror_test_item = "VUerror"
    IVerror_test_item = "IVerror"
    VVerror_test_item = "VVerror"
    IWerror_test_item = "IWerror"
    VWerror_test_item = "VWerror"

    # 从测试表格读取规格参数
    row_VAZ0 = df_test_data[df_test_data["测试项"] == VAZ0_test_item]
    row_VAZ1 = df_test_data[df_test_data["测试项"] == VAZ1_test_item]
    row_VAZ2 = df_test_data[df_test_data["测试项"] == VAZ2_test_item]
    row_IUerror = df_test_data[df_test_data["测试项"] == IUerror_test_item]
    row_VUerror = df_test_data[df_test_data["测试项"] == VUerror_test_item]
    row_IVerror = df_test_data[df_test_data["测试项"] == IVerror_test_item]
    row_VVerror = df_test_data[df_test_data["测试项"] == VVerror_test_item]
    row_IWerror = df_test_data[df_test_data["测试项"] == IWerror_test_item]
    row_VWerror = df_test_data[df_test_data["测试项"] == VWerror_test_item]

    if row_VAZ0.empty or row_VAZ1.empty or row_VAZ2.empty or row_IUerror.empty \
        or row_VUerror.empty or row_IVerror.empty or row_VVerror.empty or \
            row_IWerror.empty or row_VWerror.empty:
        print(f"未找到测试项")
        return

    try:
        VAZ0_min = float(row_VAZ0["Min."].values[0]) * 1000
        VAZ0_max = float(row_VAZ0["Max."].values[0]) * 1000
        VAZ1_min = float(row_VAZ1["Min."].values[0]) * 1000
        VAZ1_max = float(row_VAZ1["Max."].values[0]) * 1000
        VAZ2_max = float(row_VAZ2["Max."].values[0]) * 1000
        VAZ2_min = float(row_VAZ2["Min."].values[0]) * 1000
        IUerror_max = float(row_IUerror["Max."].values[0])
        VUerror_max = float(row_VUerror["Max."].values[0])
        IVerror_max = float(row_IVerror["Max."].values[0])
        VVerror_max = float(row_VVerror["Max."].values[0])
        IWerror_max = float(row_IWerror["Max."].values[0])
        VWerror_max = float(row_VWerror["Max."].values[0])


    except (IndexError, ValueError) as e:
        print(f"❌ PGA+ADC/SH+PGA测试参数读取失败: {e}")
        return

    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register)
        Pad2_ana_en = test_item.get_32bit_reg_address("Pad2_ana_en", df_register)
        Pad3_ana_en = test_item.get_32bit_reg_address("Pad3_ana_en", df_register)
        RW_ADC_CLK = test_item.get_32bit_reg_address("RW_ADC_CLK", df_register)
        RW_ADC_CTRL_EN = test_item.get_32bit_reg_address("RW_ADC_CTRL_EN", df_register)
        RW_VREF_EN = test_item.get_32bit_reg_address("RW_VREF_EN", df_register)
        RW_VREF_SEL = test_item.get_32bit_reg_address("RW_VREF_SEL", df_register)
        RW_VREF_TRIM = test_item.get_32bit_reg_address("RW_VREF_TRIM", df_register)
        RW_TEST1_EN = test_item.get_32bit_reg_address("RW_TEST1_EN", df_register)
        RW_TEST1_SEL = test_item.get_32bit_reg_address("RW_TEST1_SEL", df_register)
        RW_AMP_EN = test_item.get_32bit_reg_address("RW_AMP_EN", df_register)
        RW_PGA_EN = test_item.get_32bit_reg_address("RW_PGA_EN", df_register)
        RW_PGA_GAIN = test_item.get_32bit_reg_address("RW_PGA_GAIN", df_register)
        cfg_I_sampl_en = test_item.get_32bit_reg_address("cfg_I_sampl_en", df_register)
        CH10EN = test_item.get_32bit_reg_address("CH10EN", df_register)
        RW_ADC_EN = test_item.get_32bit_reg_address("RW_ADC_EN", df_register)
        ADCBSY = test_item.get_32bit_reg_address("ADCBSY", df_register)
        ADC_SCYC = test_item.get_32bit_reg_address("ADC_SCYC", df_register)
        cfg_az_en = test_item.get_32bit_reg_address("cfg_az_en", df_register)
        ana_az0_dat = test_item.get_32bit_reg_address("ana_az0_dat", df_register)
        ana_az1_dat = test_item.get_32bit_reg_address("ana_az1_dat", df_register)
        ana_az2_dat = test_item.get_32bit_reg_address("ana_az2_dat", df_register)

        ADC4_DR = test_item.get_32bit_reg_address("ADC4_DR", df_register)
        ADC0_DR = test_item.get_32bit_reg_address("ADC0_DR", df_register)
        ADC1_DR = test_item.get_32bit_reg_address("ADC1_DR", df_register)

    except Exception as e:
        print(f"❌ PGA+ADC/SH+PGA寄存器地址获取失败: {e}")
        return False
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    time.sleep(1)
    
    instr.swd_modify_register(Pad1_ana_en, 0xff, 8, 15)
    instr.swd_modify_register(Pad2_ana_en, 0xff, 16, 23)
    instr.swd_modify_register(Pad3_ana_en, 0xff, 24, 31)
    instr.swd_modify_register(RW_VREF_EN, 1, 8, 8)
    instr.swd_modify_register(RW_VREF_SEL, 0x03, 0, 1)
    instr.swd_modify_register(RW_VREF_TRIM, global_use["VREF_TRIM"], 23, 23)
    instr.swd_modify_register(RW_AMP_EN, 0x07, 3, 5)
    instr.swd_modify_register(RW_PGA_EN, 0x07, 11, 13)
    instr.swd_modify_register(RW_PGA_GAIN, 0x15, 8, 13)
    instr.swd_modify_register(RW_ADC_CLK, 0x01, 20, 20)
    instr.swd_modify_register(RW_ADC_CTRL_EN, 0x01, 7, 7)
    instr.swd_modify_register(cfg_I_sampl_en, 0x00, 14, 14)
    instr.swd_modify_register(RW_ADC_EN, 0x01, 1, 1)
    instr.swd_modify_register(CH10EN, 0x13, 0, 11)
    time.sleep(0.1)

    print("请将万用表接入P20引脚, 按回车继续...")
    input()
    instr.swd_modify_register(cfg_az_en, 0x01, 16, 16)
    time.sleep(0.000002)
    instr.swd_modify_register(RW_TEST1_EN, 0x01, 16, 16)

    #第一项测试项: VAZ0 测试P20输出电压
    print("请在P13、P14分别接入0V / -0.1V电压后, 按回车继续...")
    input()
    instr.swd_modify_register(RW_TEST1_SEL, 0x0C, 19, 22)
    instr.swd_modify_register(RW_TEST1_EN, 0x0, 16, 16)
    time.sleep(0.1)
    instr.swd_modify_register(ADCBSY, 0x01, 6, 6)
    ana_az0_data = instr.swd_read_register(ana_az0_dat) & 0xfff
    instr.swd_modify_register(RW_TEST1_EN, 0x1, 16, 16)
    time.sleep(0.1)
    print("VAZ0:读取P20输出电压")
    VAZ0 = instr.dmm_measure_single_voltage()
    if VAZ0 >= VAZ0_min and VAZ0 <= VAZ0_max:
        print("VAZ0 测试项通过")
    else:
        print("VAZ0 测试项不通过")
    """保存结果"""

    #第二项测试项: VAZ1 测试P20输出电压
    print("请在P16、P17分别接入0V / -0.1V电压后, 按回车继续...")
    input()
    instr.swd_modify_register(RW_TEST1_SEL, 0x0D, 19, 22)
    instr.swd_modify_register(RW_TEST1_EN, 0x0, 16, 16)
    time.sleep(0.1)
    instr.swd_modify_register(ADCBSY, 0x01, 6, 6)
    ana_az1_data = instr.swd_read_register(ana_az1_dat) & 0xfff
    instr.swd_modify_register(RW_TEST1_EN, 0x1, 16, 16)
    time.sleep(0.1)
    print("VAZ1:读取P20输出电压")
    VAZ1 = instr.dmm_measure_single_voltage()
    if VAZ1 >= VAZ1_min and VAZ1 <= VAZ1_max:
        print("VAZ1 测试项通过")
    else:
        print("VAZ1 测试项不通过")
    """保存结果"""


    #第三项测试项: VAZ2 测试P20输出电压
    print("请在P31、P32分别接入0V / -0.1V电压后, 按回车继续...")
    input()
    instr.swd_modify_register(RW_TEST1_SEL, 0x0E, 19, 22)
    instr.swd_modify_register(RW_TEST1_EN, 0x0, 16, 16)
    time.sleep(0.1)
    instr.swd_modify_register(ADCBSY, 0x01, 6, 6)
    ana_az2_data = instr.swd_read_register(ana_az2_dat) & 0xfff
    instr.swd_modify_register(RW_TEST1_EN, 0x1, 16, 16)
    time.sleep(0.1)
    print("VAZ2:读取P20输出电压")
    VAZ2 = instr.dmm_measure_single_voltage()
    if VAZ2 >= VAZ2_min and VAZ2 <= VAZ2_max:
        print("VAZ2 测试项通过")
    else:
        print("VAZ2 测试项不通过")
    """保存结果"""

    
    # 后续测试项
    """这里需要提示向三组引脚输入电压"""
    print("请分别在P13、P14和P16、P17和P31、P32接入0V / -0.1V电压后, 按回车继续...")
    input()
    
def COMP_Test():
    """ 5.12 COMP 自动修调测试"""
    # 定义测试项名称
    COMP0HYS_test_item = "COMP0HYS"
    COMP1HYS_test_item = "COMP1HYS"
    COMP2HYS_test_item = "COMP2HYS"
    COMP4HYS_test_item = "COMP4HYS"
    COMP5HYS_test_item = "COMP5HYS"
    COMP0HYS2_test_item = "COMP0HYS2"
    COMP1HYS2_test_item = "COMP1HYS2"
    COMP2HYS2_test_item = "COMP2HYS2"
    
    # 从测试表格读取规格参数
    row_COMP0HYS = df_test_data[df_test_data["测试项"] == COMP0HYS_test_item]
    row_COMP1HYS = df_test_data[df_test_data["测试项"] == COMP1HYS_test_item]
    row_COMP2HYS = df_test_data[df_test_data["测试项"] == COMP2HYS_test_item]
    row_COMP4HYS = df_test_data[df_test_data["测试项"] == COMP4HYS_test_item]
    row_COMP5HYS = df_test_data[df_test_data["测试项"] == COMP5HYS_test_item]
    row_COMP0HYS2 = df_test_data[df_test_data["测试项"] == COMP0HYS2_test_item]
    row_COMP1HYS2 = df_test_data[df_test_data["测试项"] == COMP1HYS2_test_item]
    row_COMP2HYS2 = df_test_data[df_test_data["测试项"] == COMP2HYS2_test_item]
    
    if row_COMP0HYS.empty and row_COMP1HYS.empty \
        and row_COMP2HYS.empty and row_COMP4HYS.empty \
        and row_COMP5HYS.empty  and row_COMP0HYS2.empty \
        and row_COMP1HYS2.empty  and row_COMP2HYS2.empty:
        print(f"未找到测试项")
        return
    
    try:
        # 读取COMPx规格
        COMP0HYS_min = float(row_COMP0HYS["Min."].values[0])
        COMP0HYS_max = float(row_COMP0HYS["Max."].values[0])
        COMP1HYS_min = float(row_COMP1HYS["Min."].values[0])
        COMP1HYS_max = float(row_COMP1HYS["Max."].values[0])
        COMP2HYS_max = float(row_COMP2HYS["Max."].values[0])
        COMP2HYS_min = float(row_COMP2HYS["Min."].values[0])
        COMP4HYS_max = float(row_COMP4HYS["Max."].values[0])
        COMP4HYS_min = float(row_COMP4HYS["Min."].values[0])
        COMP5HYS_max = float(row_COMP5HYS["Max."].values[0])
        COMP5HYS_min = float(row_COMP5HYS["Min."].values[0])
        COMP0HYS2_max = float(row_COMP0HYS2["Max."].values[0])
        COMP0HYS2_min = float(row_COMP0HYS2["Min."].values[0])
        COMP1HYS2_max = float(row_COMP1HYS2["Max."].values[0])
        COMP1HYS2_min = float(row_COMP1HYS2["Min."].values[0])
        COMP2HYS2_max = float(row_COMP2HYS2["Max."].values[0])
        COMP2HYS2_min = float(row_COMP2HYS2["Min."].values[0])

    except (IndexError, ValueError) as e:
        print(f"❌ VCC_S测试参数读取失败: {e}")
        return
    
    # 获取寄存器地址
    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register)
        cmp0en = test_item.get_32bit_reg_address("cmp0en", df_register)
        cmp4en = test_item.get_32bit_reg_address("cmp4en", df_register)
        Cmp5en = test_item.get_32bit_reg_address("Cmp5en", df_register)
        cmp0mod = test_item.get_32bit_reg_address("cmp0mod", df_register)
        cmp0hys = test_item.get_32bit_reg_address("cmp0hys", df_register)
        cmp1hys = test_item.get_32bit_reg_address("cmp1hys", df_register)
        Cmp2hys = test_item.get_32bit_reg_address("Cmp2hys", df_register)
        RW_TEST0_SEL = test_item.get_32bit_reg_address("RW_TEST0_SEL", df_register)
        RW_TEST0_EN = test_item.get_32bit_reg_address("RW_TEST0_EN", df_register)
        cmpsel = test_item.get_32bit_reg_address("cmpsel", df_register)
        cmp5out = test_item.get_32bit_reg_address("cmp5out", df_register)

    except Exception as e:
        print(f"❌ VCC_S寄存器地址获取失败: {e}")
        return False
    
    # 开始测试前配置 COMP 使能和时钟使能
    instr.swd_modify_register(0x40005000, 0x1, 6, 6)    # RW_AMP_CMP_EN
    instr.swd_modify_register(0x40002408, 0x1, 31, 31)  # cmp_CLK_EN
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    
    time.sleep(1)

    instr.swd_modify_register(Pad1_ana_en, 0xffffff, 8, 31)
    instr.swd_modify_register(cmp0en, 0x01, 0, 0)
    instr.swd_modify_register(cmp4en, 0x01, 7, 7)
    instr.swd_modify_register(Cmp5en, 0x01, 3, 3)
    instr.swd_modify_register(cmp0mod, 0x02, 5, 6)

    # ====================================================================
    # COMP 测试选择: 输入01245选择对应COMP, 默认全部执行
    # ====================================================================
    print("\n请选择需要测试的COMP (输入编号如 01245, 直接回车默认全部执行):")
    comp_sel = input("COMP选择: ").strip()
    if comp_sel == "":
        comp_sel = "01245"   # 默认全部
    selected = set(comp_sel)

    # ====================================================================
    # COMP0 测试: 记录P20电压步进过程中P13引脚电平翻转时的P20电压值
    # ====================================================================
    if "0" in selected:
        print("COMP0测试")
        print("万用表表笔接入P13引脚")
        print("请将DAC模块输出线接入P20引脚,固定2.5V电压接p21引脚 按回车继续...")
        input()
        instr.swd_modify_register(0x40005008, 0x01, 16, 18)
        instr.swd_modify_register(RW_TEST0_SEL, 0x01, 15, 15)
        print("\n请选择CMP0迟滞电压:")
        print("  0: 无迟滞")
        print("  1: ±2.5mV")
        print("  2: -5mV")
        print("  3: ±5mV")
        print("  4: +5mV")
        print("  5: -10mV")
        print("  6: +10mV")
        print("  7: ±10mV")
        hys_choice = input("请输入迟滞选项 (0-7): ").strip()
        try:
            hys_val = int(hys_choice) & 0x7
        except ValueError:
            print(f"输入无效: {hys_choice}, 使用默认+5mV(选项4)")
            hys_val = 4
        instr.swd_modify_register(0x40004404, hys_val, 0, 2)
        time.sleep(0.1)

        for i in range(2490, 2511):
            instr.serial_send(f"spi_cmd_set_val mV {i}\r\n")
            time.sleep(0.2)
            vol_test = instr.dmm_measure_single_voltage()
            print(f"  DAC电压={i}mV, 电压表读数={vol_test:.2f}mV")
            if vol_test >= 4000 :
                COMP0H = i
                break
        for i in range(2511, 2479, -1):
            instr.serial_send(f"spi_cmd_set_val mV {i}\r\n")
            time.sleep(0.2)
            vol_test = instr.dmm_measure_single_voltage()
            print(f"  DAC电压={i}mV, 电压表读数={vol_test:.2f}mV")
            if vol_test < 1000 :
                COMP0L = i
                break
        COMP0HYS = COMP0H - COMP0L
        print(f"COMP0HYS: {COMP0HYS}")
        if COMP0HYS >= COMP0HYS_min and COMP0HYS <= COMP0HYS_max:
            print("COMP0HYS test pass")
        else:
            print("COMP0HYS test no pass")
        """保存结果"""

    if "1" in selected:
        print("COMP1测试")
        print("万用表表笔接入P13引脚")
        print("请将DAC模块输出线接入P22引脚,固定2.5V电压接p23引脚 按回车继续...")
        input()

        instr.swd_modify_register(RW_TEST0_SEL, 0x02, 16, 18)
        instr.swd_modify_register(RW_TEST0_SEL, 0x01, 15, 15)
        print("\n请选择CMP1迟滞电压:")
        print("  0: 无迟滞")
        print("  1: ±2.5mV")
        print("  2: -5mV")
        print("  3: ±5mV")
        print("  4: +5mV")
        print("  5: -10mV")
        print("  6: +10mV")
        print("  7: ±10mV")
        hys_choice = input("请输入迟滞选项 (0-7): ").strip()
        try:
            hys_val = int(hys_choice) & 0x7
        except ValueError:
            print(f"输入无效: {hys_choice}, 使用默认+5mV(选项4)")
            hys_val = 4
        instr.swd_modify_register(0x40004404, hys_val, 7, 9)
        time.sleep(0.1)

        for i in range(2490, 2511):
            instr.serial_send(f"spi_cmd_set_val mV {i}\r\n")
            time.sleep(0.2)
            vol_test = instr.dmm_measure_single_voltage()
            print(f"  DAC电压={i}mV, 电压表读数={vol_test:.2f}mV")
            if vol_test >= 4000 :
                COMP1H = i
                break
        for i in range(2511, 2489, -1):
            instr.serial_send(f"spi_cmd_set_val mV {i}\r\n")
            time.sleep(0.2)
            vol_test = instr.dmm_measure_single_voltage()
            print(f"  DAC电压={i}mV, 电压表读数={vol_test:.2f}mV")
            if vol_test < 1000 :
                COMP1L = i
                break
        COMP1HYS = COMP1H - COMP1L
        print(f"COMP1HYS: {COMP1HYS}")
        if COMP1HYS >= COMP1HYS_min and COMP1HYS <= COMP1HYS_max:
            print("COMP1HYS test pass")
        else:
            print("COMP1HYS test no pass")
        """保存结果"""

    if "2" in selected:
        print("COMP2测试")
        print("万用表表笔接入P13引脚")
        print("请将DAC模块输出线接入P24引脚,固定2.5V电压接p25引脚 按回车继续...")
        input()
        instr.swd_modify_register(RW_TEST0_SEL, 0x03, 16, 18)
        instr.swd_modify_register(RW_TEST0_SEL, 0x01, 15, 15)
        print("\n请选择CMP2迟滞电压:")
        print("  0: 无迟滞")
        print("  1: ±2.5mV")
        print("  2: -5mV")
        print("  3: ±5mV")
        print("  4: +5mV")
        print("  5: -10mV")
        print("  6: +10mV")
        print("  7: ±10mV")
        hys_choice = input("请输入迟滞选项 (0-7): ").strip()
        try:
            hys_val = int(hys_choice) & 0x7
        except ValueError:
            print(f"输入无效: {hys_choice}, 使用默认+5mV(选项4)")
            hys_val = 4
        instr.swd_modify_register(0x40004404, hys_val, 10, 12)
        time.sleep(0.1)

        for i in range(2490, 2511):
            instr.serial_send(f"spi_cmd_set_val mV {i}\r\n")
            time.sleep(0.2)
            vol_test = instr.dmm_measure_single_voltage()
            print(f"  DAC电压={i}mV, 电压表读数={vol_test:.2f}mV")
            if vol_test >= 4000 :
                COMP2H = i
                break
        for i in range(2511, 2489, -1):
            instr.serial_send(f"spi_cmd_set_val mV {i}\r\n")
            time.sleep(0.2)
            vol_test = instr.dmm_measure_single_voltage()
            print(f"  DAC电压={i}mV, 电压表读数={vol_test:.2f}mV")
            if vol_test < 1000 :
                COMP2L = i
                break
        COMP2HYS = COMP2H - COMP2L
        print(f"COMP2HYS: {COMP2HYS}")
        if COMP2HYS >= COMP2HYS_min and COMP2HYS <= COMP2HYS_max:
            print("COMP2HYS test pass")
        else:
            print("COMP2HYS test no pass")
        """保存结果"""

    if "4" in selected:
        print("COMP4测试")
        print("无电压输出, 通过寄存器读取比较结果")
        print("请将DAC模块输出线接入P37引脚, 固定2.5V电压接P15引脚 按回车继续...")
        input()
        instr.swd_modify_register(RW_TEST0_SEL, 0x01, 15, 15)
        instr.swd_modify_register(cmpsel, 0x03, 0, 2)
        # 测试开始前清除比较结果位
        instr.swd_modify_register(0x40004410, 0x0, 7, 7)  # 清除cmp4out位

        for i in range(2450, 2551):
            instr.serial_send(f"spi_cmd_set_val mV {i}\r\n")
            time.sleep(0.2)
            cmp4_reg = instr.swd_read_register(0x40004410)
            cmp4_bit = (cmp4_reg >> 7) & 0x1
            print(f"  DAC电压={i}mV, 寄存器值={cmp4_bit}")
            if cmp4_bit == 1:
                COMP4H = i
                break
        for i in range(2551, 2449, -1):
            instr.serial_send(f"spi_cmd_set_val mV {i}\r\n")
            time.sleep(0.2)
            cmp4_reg = instr.swd_read_register(0x40004410)
            cmp4_bit = (cmp4_reg >> 7) & 0x1
            print(f"  DAC电压={i}mV, 寄存器值={cmp4_bit}")
            if cmp4_bit == 0:
                COMP4L = i
                break
        COMP4HYS = COMP4H - COMP4L
        print(f"COMP4HYS: {COMP4HYS}")
        if COMP4HYS >= COMP4HYS_min and COMP4HYS <= COMP4HYS_max:
            print("COMP4HYS test pass")
        else:
            print("COMP4HYS test no pass")
        """保存结果"""

    if "5" in selected:
        print("COMP5测试")
        print("无电压输出, 通过寄存器读取比较结果")
        print("请将DAC模块输出线接入P26引脚,固定2.5V电压接P36引脚 按回车继续...")
        input()
        instr.swd_modify_register(RW_TEST0_SEL, 0x01, 15, 15)
        # 测试开始前清除比较结果位
        instr.swd_modify_register(0x40004410, 0x0, 6, 6)  # 清除cmp5out位

        for i in range(2450, 2551):
            instr.serial_send(f"spi_cmd_set_val mV {i}\r\n")
            time.sleep(0.2)
            cmp5_reg = instr.swd_read_register(0x40004410)
            cmp5_bit = (cmp5_reg >> 6) & 0x1
            print(f"  DAC电压={i}mV, 寄存器值={cmp5_bit}")
            if cmp5_bit == 1:
                COMP5H = i
                break
        for i in range(2551, 2449, -1):
            instr.serial_send(f"spi_cmd_set_val mV {i}\r\n")
            time.sleep(0.2)
            cmp5_reg = instr.swd_read_register(0x40004410)
            cmp5_bit = (cmp5_reg >> 6) & 0x1
            print(f"  DAC电压={i}mV, 寄存器值={cmp5_bit}")
            if cmp5_bit == 0:
                COMP5L = i
                break
        COMP5HYS = COMP5H - COMP5L
        print(f"COMP5HYS: {COMP5HYS}")
        if COMP5HYS >= COMP5HYS_min and COMP5HYS <= COMP5HYS_max:
            print("COMP5HYS test pass")
        else:
            print("COMP5HYS test no pass")
        """保存结果"""



"""以下函数只是导入测试项和寄存器地址，具体测试项步骤还需要编写"""

def BEMF_Test():
    """5.13 BEMF 电阻自动修调测试"""
    # 定义测试项名称
    RES_BEMF_test_item = "RES_BEMF"
    RES_BEMF_ERROR_test_item = "RES_BEMF_ERROR"
    
    # 从测试表格读取规格参数
    row_RES_BEMF = df_test_data[df_test_data["测试项"] == RES_BEMF_test_item]
    row_RES_BEMF_ERROR = df_test_data[df_test_data["测试项"] == RES_BEMF_ERROR_test_item]
    
    if row_RES_BEMF.empty or row_RES_BEMF_ERROR.empty:
        print(f"未找到测试项: {row_RES_BEMF} 或 {row_RES_BEMF_ERROR}")
        return
    
    try:
        RES_BEMF_min = float(row_RES_BEMF["Min."].values[0])
        RES_BEMF_max = float(row_RES_BEMF["Max."].values[0])
        RES_BEMF_unit = str(row_RES_BEMF["Unit"].values[0])
        RES_BEMF_ERROR_min = float(row_RES_BEMF_ERROR["Min."].values[0])
        RES_BEMF_ERROR_max = float(row_RES_BEMF_ERROR["Max."].values[0])
        RES_BEMF_ERROR_unit = str(row_RES_BEMF_ERROR["Unit"].values[0])
    except (IndexError, ValueError) as e:
        print(f"❌ VCC_S测试参数读取失败: {e}")
        return
    
    # 获取寄存器地址
    try:
        Pad2_ana_en = test_item.get_32bit_reg_address("Pad2_ana_en", df_register)
        cmp0mod = test_item.get_32bit_reg_address("cmp0mod", df_register)
    except Exception as e:
        print(f"❌ VCC_S寄存器地址获取失败: {e}")
        return False
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    time.sleep(1)

def IO_Test():
    """5.14 IO 自动修调测试"""
    IO_OH_test_item = "IO-OH"
    IO_OL_test_item = "IO-OL"
    IO_IH_test_item = "IO-IH"
    IO_IL_test_item = "IO-IL"
    IO_PU_RES_test_item = "IO_PU_RES"
    IO_PD_RES_test_item = "IO_PD_RES"

    row_IO_OH = df_test_data[df_test_data["测试项"] ==      IO_OH_test_item]
    row_IO_OL = df_test_data[df_test_data["测试项"] ==      IO_OL_test_item]
    row_IO_IH = df_test_data[df_test_data["测试项"] ==      IO_IH_test_item]
    row_IO_IL = df_test_data[df_test_data["测试项"] ==      IO_IL_test_item]
    row_IO_PU_RES = df_test_data[df_test_data["测试项"] ==  IO_PU_RES_test_item]
    row_IO_PD_RES = df_test_data[df_test_data["测试项"] ==  IO_PD_RES_test_item]

    try:
        # 读取COMP规格
        IO_OH_min = float(row_IO_OH["Min."].values[0])
        IO_OH_max = float(row_IO_OH["Max."].values[0])
        IO_OL_min = float(row_IO_OL["Min."].values[0])
        IO_OL_max = float(row_IO_OL["Max."].values[0])
        IO_IH_min = float(row_IO_IH["Min."].values[0])
        IO_IL_max = float(row_IO_IL["Max."].values[0])
        IO_PU_RES_max = float(row_IO_PU_RES["Max."].values[0])
        IO_PU_RES_min = float(row_IO_PU_RES["Min."].values[0])
        IO_PD_RES_max = float(row_IO_PD_RES["Max."].values[0])
        IO_PD_RES_min = float(row_IO_PD_RES["Min."].values[0])

    except (IndexError, ValueError) as e:
        print(f"❌ VCC_S测试参数读取失败: {e}")
        return

    # 获取寄存器地址
    try:
        Pad1_ana_en = test_item.get_32bit_reg_address("Pad1_ana_en", df_register)
        Pad4_ana_en_2 = test_item.get_32bit_reg_address("Pad4_ana_en_2", df_register)
        pad_rstn_en = test_item.get_32bit_reg_address("pad_rstn_en", df_register)
        Gpio4_out_en = test_item.get_32bit_reg_address("Gpio4_out_en", df_register)
        Gpio_out_en = test_item.get_32bit_reg_address("Gpio_out_en", df_register)
        Gpio_out = test_item.get_32bit_reg_address("Gpio_out", df_register)
        Gpio4_out = test_item.get_32bit_reg_address("Gpio4_out", df_register)
        Gpio_in = test_item.get_32bit_reg_address("Gpio_in", df_register)
        Gpio4_in = test_item.get_32bit_reg_address("Gpio4_in", df_register)
        Pad_pu = test_item.get_32bit_reg_address("Pad_pu", df_register)
        Pad4_pu = test_item.get_32bit_reg_address("Pad4_pu", df_register)
        cmp5out = test_item.get_32bit_reg_address("cmp5out", df_register)
        Pad1_pd = test_item.get_32bit_reg_address("Pad1_pd", df_register)
        Pad4_pd = test_item.get_32bit_reg_address("Pad4_pd", df_register)

    except Exception as e:
        print(f"❌ VCC_S寄存器地址获取失败: {e}")
        return False
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    time.sleep(1)
    
    instr.swd_modify_register(Pad1_ana_en, 0x0, 8, 31)
    instr.swd_modify_register(Pad4_ana_en_2, 0x0, 26, 26)
    instr.swd_modify_register(pad_rstn_en, 0x0, 27, 27)
    instr.swd_modify_register(Gpio4_out_en, 0x3f, 16, 21)
    instr.swd_modify_register(Gpio_out_en, 0xfffffffc, 0, 31)
    instr.swd_modify_register(Gpio_out, 0xfffffffc, 0, 31)
    instr.swd_modify_register(Gpio4_out, 0x3f, 8, 13)
    
def Program_Test():
    """5.15 程序自动修调测试 (需要配合MCU程序测试) """

def Chip_MTP_Write_Read_Test():
    """5.16 芯片MTP读写操作"""
    # 获取寄存器地址( 与项目统一寄存器名一致) 
    try:
        Ecc_en = test_item.get_32bit_reg_address("Ecc_en", df_register)
        mtp_remap = test_item.get_32bit_reg_address("mtp_remap", df_register)
        mtp_wdat = test_item.get_32bit_reg_address("mtp_wdat", df_register)
        Wo_er = test_item.get_32bit_reg_address("Wo_er", df_register)
        Wo_per = test_item.get_32bit_reg_address("Wo_per", df_register)
        AD = test_item.get_32bit_reg_address("AD", df_register)
    except Exception as e:
        print(f"❌ Chip_MTP寄存器地址获取失败: {e}")
        return False
    
    #instr.serial_send("spi_cmd_set_val mV 5000\r\n")
    time.sleep(1)

# 串口通信演示函数, 该函数不参与自动化测试，仅作为示例展示如何在测试中使用串口通信功能
def Serial_Comm_Demo():
    """串口通信演示函数, 展示如何在测试中使用串口"""
    print("\n🚀 串口通信演示开始...")

    # 使用全局串口实例发送数据
    result = instr.serial_send("AT+TEST\r\n")
    if result > 0:
        print(f"✅ 串口数据发送成功, 发送字节数: {result}")

        # 接收响应
        time.sleep(0.1)
        response = instr.serial_receive()
        if response:
            print(f"📨 收到串口响应: {response.decode('utf-8', 'ignore')}")
        else:
            print("⚠️  未收到串口响应")
    else:
        print("❌ 串口数据发送失败")

    print("🎉 串口通信演示结束")
